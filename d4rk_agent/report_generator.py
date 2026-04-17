"""
SIRIAN REPORT GENERATOR — 전문 리포트 자동 생성
주제를 받아서 PPTX/XLSX 리포트로 만들어주기
웹에서 이미지 수집 + 구체적 설명 포함
"""
import os, json, logging, re, requests, tempfile
from datetime import datetime
from utils import ask_qwen, strip_chinese

log = logging.getLogger("report_gen")

OUTPUT_DIR = "C:/Users/gohun/Desktop/sirian/sirian_space/reports"
AGENT_DIR  = "C:/Users/gohun/Desktop/sirian/d4rk_agent"

class ReportGenerator:
    def __init__(self):
        os.makedirs(OUTPUT_DIR, exist_ok=True)

    # ─── 메인 진입점 ───
    def generate(self, topic: str, format: str = "pptx",
                 level: str = "초보자") -> dict:
        """
        topic: 주제 (예: "SQL Injection")
        format: pptx / xlsx / both
        level: 초보자 / 중급 / 전문가
        """
        log.info(f"리포트 생성 시작: {topic} ({format}, {level})")

        # 1. 콘텐츠 생성
        content = self._research_topic(topic, level)

        # 2. 이미지 수집
        images = self._collect_images(topic)

        # 3. 파일 생성
        results = {}
        ts = datetime.now().strftime("%Y%m%d_%H%M")
        safe_topic = re.sub(r'[^\w가-힣]', '_', topic)[:20]

        if format in ("pptx", "both"):
            pptx_path = f"{OUTPUT_DIR}/{ts}_{safe_topic}.pptx"
            self._create_pptx(topic, content, images, pptx_path, level)
            results["pptx"] = pptx_path

        if format in ("xlsx", "both"):
            xlsx_path = f"{OUTPUT_DIR}/{ts}_{safe_topic}.xlsx"
            self._create_xlsx(topic, content, xlsx_path)
            results["xlsx"] = xlsx_path

        log.info(f"리포트 완료: {results}")
        return results

    # ─── 콘텐츠 생성 ───
    def _research_topic(self, topic: str, level: str) -> dict:
        """주제 연구 — 웹 검색 + qwen 정리"""
        # 웹 검색
        web_results = []
        try:
            from tools import tools
            web_results = tools.web_search(topic, max_results=5)
        except: pass

        web_ctx = "\n".join([
            f"- {r.get('title','')}: {r.get('snippet','')}"
            for r in web_results[:3]
        ]) if web_results else ""

        level_guide = {
            "초보자": "전문 용어 없이, 쉬운 예시와 비유 사용",
            "중급":   "기술적 내용 포함, 실습 예제 포함",
            "전문가": "심층 기술 분석, CVE/코드 포함"
        }.get(level, "쉬운 설명")

        # 섹션별 내용 생성
        sections = {}

        # 개요
        prompt = (
            f"주제: {topic}\n참고: {web_ctx[:500]}\n대상: {level_guide}\n\n"
            "1. 개요 (무엇인가, 왜 중요한가) — 300자 이내. 한국어로."
        )
        sections["overview"] = ask_qwen(prompt, max_tokens=300, temperature=0.6)

        # 핵심 내용 (구조화)
        prompt2 = (
            f"주제: {topic}\n참고: {web_ctx[:500]}\n대상: {level_guide}\n\n"
            "핵심 내용 5가지를 JSON으로:\n"
            '[{"title":"...", "content":"...", "example":"..."}]\n'
            "각 항목은 100자 이내. 한국어로."
        )
        raw = ask_qwen(prompt2, max_tokens=800, temperature=0.5)
        try:
            match = re.search(r'\[.*?\]', raw, re.DOTALL)
            sections["keypoints"] = json.loads(match.group()) if match else []
        except:
            sections["keypoints"] = []

        # 취약점/위험 (보안 주제인 경우)
        if any(k in topic.lower() for k in ['injection','xss','rce','sqli','rop','exploit','취약','해킹','보안']):
            prompt3 = (
                f"주제: {topic}\n대상: {level_guide}\n\n"
                "취약한 코드 예시와 안전한 코드 예시를 각각:\n"
                '{"vulnerable": "코드 예시", "safe": "코드 예시", "reason": "이유"}\n'
                "한국어로."
            )
            raw3 = ask_qwen(prompt3, max_tokens=400, temperature=0.4)
            try:
                match3 = re.search(r'\{.*?\}', raw3, re.DOTALL)
                sections["code_example"] = json.loads(match3.group()) if match3 else {}
            except:
                sections["code_example"] = {}

        # 예방/대응 방법
        prompt4 = (
            f"주제: {topic}\n대상: {level_guide}\n\n"
            "예방/대응 방법 4가지. JSON:\n"
            '[{"method":"...", "description":"...", "difficulty":"쉬움/보통/어려움"}]\n'
            "한국어로."
        )
        raw4 = ask_qwen(prompt4, max_tokens=400, temperature=0.5)
        try:
            match4 = re.search(r'\[.*?\]', raw4, re.DOTALL)
            sections["countermeasures"] = json.loads(match4.group()) if match4 else []
        except:
            sections["countermeasures"] = []

        # 실제 사례
        prompt5 = (
            f"주제: {topic}\n참고: {web_ctx[:300]}\n\n"
            "실제 사례나 유명한 사건 2~3개. 한국어로 각 100자 이내."
        )
        sections["cases"] = ask_qwen(prompt5, max_tokens=300, temperature=0.6)

        return sections

    # ─── 이미지 수집 ───
    def _collect_images(self, topic: str) -> list:
        """웹에서 관련 이미지 URL 수집"""
        images = []
        try:
            from tools import tools
            results = tools.web_search(f"{topic} diagram infographic", max_results=5)
            for r in results[:3]:
                url = r.get("image_url") or r.get("thumbnail")
                if url and url.startswith("http"):
                    images.append(url)
        except: pass

        # 다운로드
        downloaded = []
        for url in images[:3]:
            try:
                resp = requests.get(url, timeout=5)
                if resp.status_code == 200 and "image" in resp.headers.get("content-type",""):
                    ext = ".jpg" if "jpeg" in resp.headers.get("content-type","") else ".png"
                    with tempfile.NamedTemporaryFile(suffix=ext, delete=False) as f:
                        f.write(resp.content)
                        downloaded.append(f.name)
            except: pass

        return downloaded

    # ─── PPTX 생성 ───
    def _create_pptx(self, topic: str, content: dict, images: list,
                     output_path: str, level: str):
        """pptxgenjs로 전문 슬라이드 생성"""
        try:
            import subprocess, json as _json

            # 슬라이드 데이터 구성
            slides_data = self._build_slides_data(topic, content, images, level)

            # Node.js 스크립트
            script = self._build_pptxgenjs_script(slides_data, output_path)
            script_file = f"{OUTPUT_DIR}/gen_script.js"
            with open(script_file,'w',encoding='utf-8') as f:
                f.write(script)

            result = subprocess.run(
                ['node', script_file],
                capture_output=True, text=True, timeout=60
            )
            if result.returncode == 0:
                log.info(f"PPTX 생성 완료: {output_path}")
            else:
                log.error(f"PPTX 생성 실패: {result.stderr[:200]}")
        except Exception as e:
            log.error(f"PPTX 오류: {e}")

    def _build_slides_data(self, topic: str, content: dict,
                           images: list, level: str) -> list:
        slides = []
        # 1. 표지
        slides.append({
            "type": "title",
            "title": topic,
            "subtitle": f"완벽 이해 가이드 | 대상: {level} | {datetime.now().strftime('%Y.%m.%d')}",
        })
        # 2. 개요
        if content.get("overview"):
            slides.append({
                "type": "content",
                "title": "📌 개요",
                "body": content["overview"],
                "image": images[0] if images else None
            })
        # 3~7. 핵심 포인트
        for i, kp in enumerate(content.get("keypoints",[])[:5]):
            slides.append({
                "type": "keypoint",
                "title": kp.get("title",""),
                "content": kp.get("content",""),
                "example": kp.get("example",""),
                "image": images[min(i+1, len(images)-1)] if images else None
            })
        # 8. 코드 예시
        if content.get("code_example"):
            ce = content["code_example"]
            slides.append({
                "type": "code",
                "title": "⚠️ 취약 코드 vs ✅ 안전 코드",
                "vulnerable": ce.get("vulnerable",""),
                "safe": ce.get("safe",""),
                "reason": ce.get("reason","")
            })
        # 9. 대응 방법
        if content.get("countermeasures"):
            slides.append({
                "type": "countermeasures",
                "title": "🛡️ 예방 및 대응 방법",
                "items": content["countermeasures"]
            })
        # 10. 실제 사례
        if content.get("cases"):
            slides.append({
                "type": "content",
                "title": "📰 실제 사례",
                "body": content["cases"]
            })
        # 11. 마무리
        slides.append({
            "type": "end",
            "title": "감사합니다",
            "subtitle": f"{topic} 완벽 이해 가이드"
        })
        return slides

    def _build_pptxgenjs_script(self, slides: list, output_path: str) -> str:
        slides_json = json.dumps(slides, ensure_ascii=False)
        output_path_js = output_path.replace('\\','\\\\')

        return f"""
const pptxgen = require('pptxgenjs');
const prs = new pptxgen();

// 슬라이드 크기 설정
prs.layout = 'LAYOUT_WIDE';

// 색상 테마 (보안 = 다크 네이비 + 청록)
const C = {{
  bg:      '0D1B2A',
  panel:   '1B2A3B',
  accent:  '00D4FF',
  accent2: '7B2FFF',
  text:    'E8F4FD',
  dim:     '6B8FAD',
  green:   '00FF9D',
  red:     'FF2D6E',
  white:   'FFFFFF',
}};

const slides = {slides_json};

slides.forEach((s, idx) => {{
  const slide = prs.addSlide();
  slide.background = {{ color: C.bg }};

  if (s.type === 'title') {{
    // 표지
    slide.addShape(prs.ShapeType.rect, {{
      x:0, y:0, w:'100%', h:0.08,
      fill: {{ color: C.accent }}
    }});
    slide.addShape(prs.ShapeType.rect, {{
      x:0, y:'92%', w:'100%', h:0.08,
      fill: {{ color: C.accent2 }}
    }});
    slide.addText(s.title, {{
      x:1, y:2.5, w:11, h:1.5,
      fontSize:44, bold:true, color:C.accent,
      fontFace:'Arial Black', align:'center'
    }});
    slide.addText(s.subtitle, {{
      x:1, y:4.2, w:11, h:0.6,
      fontSize:16, color:C.dim, align:'center'
    }});

  }} else if (s.type === 'content') {{
    // 일반 콘텐츠
    slide.addText(s.title, {{
      x:0.5, y:0.3, w:12, h:0.7,
      fontSize:28, bold:true, color:C.accent, fontFace:'Arial'
    }});
    slide.addShape(prs.ShapeType.rect, {{
      x:0.5, y:1.1, w:11, h:0.04,
      fill: {{ color: C.accent }}
    }});
    slide.addText(s.body || '', {{
      x:0.5, y:1.3, w: s.image ? 7.5 : 12, h:5,
      fontSize:15, color:C.text, valign:'top',
      breakLine:true, lineSpacingMultiple:1.3
    }});
    if (s.image) {{
      try {{
        slide.addImage({{ path: s.image, x:8.2, y:1.3, w:4.5, h:4.5 }});
      }} catch(e) {{}}
    }}

  }} else if (s.type === 'keypoint') {{
    slide.addText(s.title, {{
      x:0.5, y:0.3, w:12, h:0.7,
      fontSize:26, bold:true, color:C.accent2
    }});
    slide.addShape(prs.ShapeType.roundRect, {{
      x:0.5, y:1.2, w:8, h:2.5,
      fill:{{ color: C.panel }}, line:{{ color:C.accent2, pt:2 }},
      rectRadius:0.1
    }});
    slide.addText(s.content || '', {{
      x:0.7, y:1.4, w:7.6, h:2.1,
      fontSize:14, color:C.text, valign:'top', breakLine:true
    }});
    if (s.example) {{
      slide.addShape(prs.ShapeType.roundRect, {{
        x:0.5, y:4.0, w:8, h:1.5,
        fill:{{ color:'1A1A2E' }}, line:{{ color:C.green, pt:1 }},
        rectRadius:0.1
      }});
      slide.addText('예시: ' + s.example, {{
        x:0.7, y:4.1, w:7.6, h:1.3,
        fontSize:12, color:C.green, fontFace:'Consolas'
      }});
    }}

  }} else if (s.type === 'code') {{
    slide.addText(s.title, {{
      x:0.5, y:0.2, w:12, h:0.6,
      fontSize:24, bold:true, color:C.red
    }});
    // 취약 코드
    slide.addShape(prs.ShapeType.rect, {{
      x:0.3, y:1.0, w:6, h:3.5,
      fill:{{ color:'2A0A0A' }}, line:{{ color:C.red, pt:2 }}
    }});
    slide.addText('❌ 취약 코드', {{
      x:0.5, y:1.0, w:5.6, h:0.4,
      fontSize:12, bold:true, color:C.red
    }});
    slide.addText(s.vulnerable || '', {{
      x:0.4, y:1.5, w:5.8, h:2.8,
      fontSize:11, color:'FF8888', fontFace:'Consolas', valign:'top'
    }});
    // 안전 코드
    slide.addShape(prs.ShapeType.rect, {{
      x:6.8, y:1.0, w:6, h:3.5,
      fill:{{ color:'0A2A0A' }}, line:{{ color:C.green, pt:2 }}
    }});
    slide.addText('✅ 안전 코드', {{
      x:7.0, y:1.0, w:5.6, h:0.4,
      fontSize:12, bold:true, color:C.green
    }});
    slide.addText(s.safe || '', {{
      x:6.9, y:1.5, w:5.8, h:2.8,
      fontSize:11, color:'88FF88', fontFace:'Consolas', valign:'top'
    }});
    // 이유
    slide.addText('💡 ' + (s.reason || ''), {{
      x:0.5, y:4.7, w:12.3, h:0.8,
      fontSize:13, color:C.dim, italic:true
    }});

  }} else if (s.type === 'countermeasures') {{
    slide.addText(s.title, {{
      x:0.5, y:0.3, w:12, h:0.7,
      fontSize:28, bold:true, color:C.green
    }});
    const items = s.items || [];
    items.forEach((item, i) => {{
      const col = i % 2;
      const row = Math.floor(i / 2);
      const colors = [C.accent, C.accent2, C.green, 'FFD600'];
      slide.addShape(prs.ShapeType.roundRect, {{
        x: col * 6.3 + 0.5, y: row * 2.5 + 1.2, w:5.8, h:2.2,
        fill:{{ color: C.panel }}, line:{{ color: colors[i], pt:2 }},
        rectRadius:0.15
      }});
      slide.addText(item.method || '', {{
        x: col*6.3+0.8, y: row*2.5+1.3, w:5.3, h:0.5,
        fontSize:14, bold:true, color: colors[i]
      }});
      slide.addText(item.description || '', {{
        x: col*6.3+0.8, y: row*2.5+1.85, w:5.2, h:1.3,
        fontSize:12, color:C.text, valign:'top'
      }});
    }});

  }} else if (s.type === 'end') {{
    slide.addShape(prs.ShapeType.rect, {{
      x:0, y:0, w:'100%', h:'100%',
      fill:{{ color: C.panel }}
    }});
    slide.addText(s.title, {{
      x:1, y:2.8, w:11, h:1.2,
      fontSize:48, bold:true, color:C.accent,
      fontFace:'Arial Black', align:'center'
    }});
    slide.addText(s.subtitle, {{
      x:1, y:4.3, w:11, h:0.6,
      fontSize:18, color:C.dim, align:'center'
    }});
  }}
}});

prs.writeFile({{ fileName: '{output_path_js}' }})
  .then(() => console.log('완료: {output_path_js}'))
  .catch(e => console.error('오류:', e));
"""

    # ─── XLSX 생성 ───
    def _create_xlsx(self, topic: str, content: dict, output_path: str):
        try:
            import openpyxl
            from openpyxl.styles import (PatternFill, Font, Alignment,
                                         Border, Side)
            from openpyxl.utils import get_column_letter

            wb = openpyxl.Workbook()

            # 색상
            BG     = "0D1B2A"
            ACCENT = "00D4FF"
            PANEL  = "1B2A3B"
            GREEN  = "00FF9D"
            RED    = "FF2D6E"
            WHITE  = "E8F4FD"
            DIM    = "6B8FAD"

            def hdr_fill(color): return PatternFill("solid", fgColor=color)
            def hdr_font(color="FFFFFF", sz=12, bold=True):
                return Font(color=color, size=sz, bold=bold)
            def center(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
            def left():   return Alignment(horizontal="left",   vertical="center", wrap_text=True)

            # ── 시트 1: 개요 ──
            ws1 = wb.active
            ws1.title = "📌 개요"
            ws1.sheet_view.showGridLines = False
            ws1.column_dimensions["A"].width = 20
            ws1.column_dimensions["B"].width = 80

            ws1["A1"] = topic + " — 완벽 가이드"
            ws1["A1"].font = Font(color=ACCENT, size=18, bold=True)
            ws1.merge_cells("A1:B1")
            ws1["A1"].alignment = center()
            ws1.row_dimensions[1].height = 40

            ws1["A3"] = "개요"
            ws1["A3"].fill = hdr_fill(ACCENT)
            ws1["A3"].font = hdr_font("0D1B2A", 12)
            ws1["A3"].alignment = center()
            ws1["B3"] = content.get("overview","")
            ws1["B3"].alignment = left()
            ws1.row_dimensions[3].height = 80

            # 핵심 포인트
            ws1["A5"] = "핵심 포인트"
            ws1["A5"].fill = hdr_fill(PANEL)
            ws1["A5"].font = hdr_font(ACCENT, 11)
            ws1["A5"].alignment = center()
            ws1["B5"] = "설명"
            ws1["B5"].fill = hdr_fill(PANEL)
            ws1["B5"].font = hdr_font(ACCENT, 11)
            ws1["B5"].alignment = center()

            for i, kp in enumerate(content.get("keypoints",[]), start=6):
                ws1[f"A{i}"] = kp.get("title","")
                ws1[f"A{i}"].fill = hdr_fill("162030")
                ws1[f"A{i}"].font = Font(color=WHITE, size=11, bold=True)
                ws1[f"A{i}"].alignment = center()
                ws1[f"B{i}"] = kp.get("content","")
                ws1[f"B{i}"].fill = hdr_fill("0F1E2D")
                ws1[f"B{i}"].font = Font(color=WHITE, size=10)
                ws1[f"B{i}"].alignment = left()
                ws1.row_dimensions[i].height = 50

            # ── 시트 2: 대응방법 ──
            ws2 = wb.create_sheet("🛡️ 대응방법")
            ws2.sheet_view.showGridLines = False
            ws2.column_dimensions["A"].width = 25
            ws2.column_dimensions["B"].width = 50
            ws2.column_dimensions["C"].width = 15

            ws2["A1"] = "예방 및 대응 방법"
            ws2["A1"].font = Font(color=GREEN, size=16, bold=True)
            ws2.merge_cells("A1:C1")
            ws2["A1"].alignment = center()

            headers = ["방법", "설명", "난이도"]
            colors  = [ACCENT, DIM, GREEN]
            for col, (h, c) in enumerate(zip(headers, colors), 1):
                cell = ws2.cell(3, col, h)
                cell.fill   = hdr_fill(PANEL)
                cell.font   = Font(color=c, size=11, bold=True)
                cell.alignment = center()

            diff_colors = {"쉬움": GREEN, "보통": "FFD600", "어려움": RED}
            for i, cm in enumerate(content.get("countermeasures",[]), start=4):
                ws2.cell(i,1, cm.get("method","")).font = Font(color=WHITE, bold=True)
                ws2.cell(i,2, cm.get("description","")).font = Font(color=WHITE)
                diff = cm.get("difficulty","보통")
                dc = ws2.cell(i,3, diff)
                dc.font = Font(color=diff_colors.get(diff, WHITE), bold=True)
                dc.alignment = center()
                for col in range(1,4):
                    ws2.cell(i,col).fill = hdr_fill("0F1E2D")
                    ws2.cell(i,col).alignment = left() if col==2 else center()
                ws2.row_dimensions[i].height = 40

            # ── 시트 3: 코드 예시 ──
            if content.get("code_example"):
                ws3 = wb.create_sheet("💻 코드 예시")
                ws3.sheet_view.showGridLines = False
                ws3.column_dimensions["A"].width = 15
                ws3.column_dimensions["B"].width = 65

                ws3["A1"] = "코드 비교"
                ws3["A1"].font = Font(color=RED, size=14, bold=True)
                ws3.merge_cells("A1:B1")

                ce = content["code_example"]
                for row, (label, val, color) in enumerate([
                    ("❌ 취약 코드", ce.get("vulnerable",""), RED),
                    ("✅ 안전 코드", ce.get("safe",""), GREEN),
                    ("💡 이유", ce.get("reason",""), ACCENT),
                ], start=3):
                    ws3.cell(row,1,label).fill = hdr_fill(PANEL)
                    ws3.cell(row,1,label).font = Font(color=color, bold=True)
                    ws3.cell(row,1,label).alignment = center()
                    ws3.cell(row,2,val).fill = hdr_fill("0A0F1A")
                    ws3.cell(row,2,val).font = Font(color=color, name="Consolas", size=10)
                    ws3.cell(row,2,val).alignment = left()
                    ws3.row_dimensions[row].height = 80

            wb.save(output_path)
            log.info(f"XLSX 생성 완료: {output_path}")
        except Exception as e:
            log.error(f"XLSX 오류: {e}")

    def notify_done(self, results: dict, topic: str):
        """완료 알림"""
        msg = f"{topic} 리포트 완성했어."
        if "pptx" in results:
            msg += " PPT도 만들었어."
        if "xlsx" in results:
            msg += " 엑셀도 만들었어."
        msg += f" {OUTPUT_DIR} 폴더에 저장했어."
        try:
            from tts_engine import tts
            tts.speak(msg, priority=True)
        except: pass
        return msg

report_generator = ReportGenerator()
